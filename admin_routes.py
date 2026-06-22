from flask import Blueprint, render_template, session, redirect, request, jsonify, flash
from db import get_db
from datetime import timedelta

admin_bp = Blueprint('admin', __name__)

# Thêm vào dòng import ở đầu file
import json

# Thêm hàm này ngay sau các import, trước @admin_bp.route đầu tiên
def td_to_str(val):
    if val is None:
        return ''
    if isinstance(val, timedelta):
        total = int(val.total_seconds())
        return f"{total//3600:02d}:{(total%3600)//60:02d}"
    return str(val)[:5]

# --- 1. TRANG DANH SÁCH & TÌM KIẾM ---
@admin_bp.route("/manage_users")
def manage_users():
    if session.get("role") != "admin":
        return redirect("/login")
    
    search_query = request.args.get('search', '').strip()
    
    db = get_db()
    cursor = db.cursor(dictionary=True)
    
    try:
        if search_query:

            query = """
                SELECT u.id, u.username, u.role, u.status,
                       s.name as student_name, s.student_code, s.class_name,
                       t.name as teacher_name, t.department
                FROM users u
                LEFT JOIN students s ON u.id = s.user_id
                LEFT JOIN teachers t ON u.id = t.user_id
                WHERE u.username LIKE %s 
                   OR s.name LIKE %s 
                   OR t.name LIKE %s 
                   OR s.student_code LIKE %s
                ORDER BY u.id DESC
            """
            like_val = f"%{search_query}%"
            cursor.execute(query, (like_val, like_val, like_val, like_val))
        else:
            # Nếu không tìm kiếm: Lấy toàn bộ danh sách
            query = """
                SELECT u.id, u.username, u.role, u.status,
                       s.name as student_name, s.student_code, s.class_name,
                       t.name as teacher_name, t.department
                FROM users u
                LEFT JOIN students s ON u.id = s.user_id
                LEFT JOIN teachers t ON u.id = t.user_id
                ORDER BY u.id DESC
            """
            cursor.execute(query)
            
        users_list = cursor.fetchall()
        
        # Trả về kết quả kèm biến 'search' để ô input giữ lại từ khóa vừa nhập
        return render_template("manage_users.html", users=users_list, search=search_query)
        
    finally:
        cursor.close()
        db.close()

# --- 2. THÊM TÀI KHOẢN MỚI ---
@admin_bp.route("/add_user", methods=["POST"])
def add_user():
    if session.get("role") != "admin": return redirect("/login")
    
    username = request.form.get("username")
    password = request.form.get("password")
    role = request.form.get("role")
    
    db = get_db()
    cursor = db.cursor()
    try:
        # Thêm vào bảng users
        cursor.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)", 
                       (username, password, role))
        user_id = cursor.lastrowid
        
        # Nếu là student, tạo sẵn 1 dòng bên bảng students để liên kết
        if role == 'student':
            student_code = request.form.get("student_code")
            cursor.execute("INSERT INTO students (user_id, student_code) VALUES (%s, %s)", (user_id, student_code))
        
        # Nếu là teacher, tạo sẵn 1 dòng bên bảng teachers
        elif role == 'teacher':
            full_name = request.form.get("full_name")
            cursor.execute("INSERT INTO teachers (user_id, name) VALUES (%s, %s)", (user_id, full_name))
            
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"Error adding user: {e}")
    finally:
        cursor.close()
        db.close()
    return redirect("/manage_users")

# --- 3. XÓA TÀI KHOẢN ---
@admin_bp.route("/delete_user/<int:user_id>")
def delete_user(user_id):
    if session.get("role") != "admin": return redirect("/login")
    
    db = get_db()
    cursor = db.cursor()
    try:
        # Xóa ở bảng users (Các bảng students/teachers nên cài ON DELETE CASCADE ở DB 
        # Nếu chưa cài thì ta phải xóa thủ công ở bảng con trước)
        cursor.execute("DELETE FROM students WHERE user_id = %s", (user_id,))
        cursor.execute("DELETE FROM teachers WHERE user_id = %s", (user_id,))
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"Error: {e}")
    finally:
        cursor.close()
        db.close()
    return redirect("/manage_users")

# --- 4. CẬP NHẬT TÀI KHOẢN (API lấy data và POST update) ---
@admin_bp.route("/edit_user", methods=["POST"])
def edit_user():
    if session.get("role") != "admin": return redirect("/login")
    
    user_id = request.form.get("user_id")
    full_name = request.form.get("full_name")
    status = request.form.get("status")
    role = request.form.get("role")
    
    db = get_db()
    cursor = db.cursor()
    try:
        # 1. Cập nhật trạng thái ở bảng users
        cursor.execute("UPDATE users SET status=%s WHERE id=%s", (status, user_id))
        
        # 2. Cập nhật chi tiết theo vai trò
        if role == 'student':
            class_name = request.form.get("class_name")
            cursor.execute("""
                UPDATE students SET name=%s, class_name=%s WHERE user_id=%s
            """, (full_name, class_name, user_id))
        elif role == 'teacher':
            department = request.form.get("department")
            # Giả sử bảng teachers có cột department (khoa)
            cursor.execute("""
                UPDATE teachers SET name=%s, department=%s WHERE user_id=%s
            """, (full_name, department, user_id))
            
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"Lỗi cập nhật: {e}")
    finally:
        cursor.close()
        db.close()
    return redirect("/manage_users")



@admin_bp.route('/manage_teachers')
def manage_teachers():
    if session.get("role") != "admin":
        return redirect("/login")

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT t.*, u.username, u.status
            FROM teachers t
            JOIN users u ON t.user_id = u.id
        """)
        teachers_list = cursor.fetchall()

        # BUG FIX: Vòng lặp chỉ ở cấp ngoài, không lồng nhau
        for teacher in teachers_list:
            cursor.execute("""
                SELECT
                    sub.id          AS subject_id,
                    sub.name        AS subject_name,
                    sub.class_name  AS subject_class,
                    sch.id          AS schedule_id,
                    sch.room,
                    sch.start_time,
                    sch.end_time,
                    sch.start_date,
                    sch.end_date,
                    sch.days_of_week,
                    COALESCE(JSON_LENGTH(sch.days_of_week), 0) AS lessons_per_week
                FROM subjects sub
                LEFT JOIN schedules sch ON sch.subject_id = sub.id
                WHERE sub.teacher_id = %s
                ORDER BY sub.id, sch.id
            """, (teacher['id'],))

            rows = cursor.fetchall()
            subjects_dict = {}

            for row in rows:
                sid = row['subject_id']
                if sid not in subjects_dict:
                    subjects_dict[sid] = {
                        'subject_id':   sid,
                        'subject_name': row['subject_name'],
                        'class_name':   row['subject_class'] or '',
                        'schedules':    []
                    }
                if row['schedule_id']:
                    days = row['days_of_week']
                    if isinstance(days, str):
                        try:
                            days = json.loads(days)
                        except Exception:
                            days = []
                    elif days is None:
                        days = []

                    subjects_dict[sid]['schedules'].append({
                        'schedule_id':      row['schedule_id'],
                        'class_name':       row['subject_class'] or '',
                        'room':             row['room'] or '',
                        'start_time':       td_to_str(row['start_time']),
                        'end_time':         td_to_str(row['end_time']),
                        'start_date':       str(row['start_date']) if row['start_date'] else '',
                        'end_date':         str(row['end_date'])   if row['end_date']   else '',
                        'days_of_week':     days,
                        'lessons_per_week': int(row['lessons_per_week'] or 0)
                    })

            teacher['teaching_subjects'] = list(subjects_dict.values())
            teacher['total_lessons'] = sum(
                sch['lessons_per_week']
                for s in teacher['teaching_subjects']
                for sch in s['schedules']
            )

        return render_template('manage_teacher.html', teachers=teachers_list)
    finally:
        cursor.close()
        db.close()



@admin_bp.route('/manage_subjects')
def manage_subjects():
    if session.get("role") != "admin":
        return redirect("/login")

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        # Lấy danh sách môn học + tên giáo viên
        cursor.execute("""
            SELECT s.id,
                   s.name,
                   s.class_name,
                   t.id AS teacher_id,
                   t.name AS teacher_name,
                   t.department
            FROM subjects s
            LEFT JOIN teachers t
                ON s.teacher_id = t.id
            ORDER BY s.id DESC
        """)
        subjects_list = cursor.fetchall()
        # Lấy danh sách lịch học + tên môn + giáo viên
        cursor.execute("""
            SELECT sch.id,
                   sch.subject_id,
                   sch.class_name,
                   sch.room,
                   TIME_FORMAT(sch.start_time,'%H:%i') AS start_time,
                   TIME_FORMAT(sch.end_time,'%H:%i') AS end_time,
                   sch.start_date,
                   sch.end_date,
                   sch.days_of_week,
                   s.name AS subject_name,
                   t.id AS teacher_id,
                   t.name AS teacher_name
            FROM schedules sch
            JOIN subjects s
                ON sch.subject_id = s.id
            LEFT JOIN teachers t
                ON s.teacher_id = t.id
            ORDER BY sch.id DESC
        """)
        schedules_list = cursor.fetchall()


        # Xử lý days_of_week từ JSON string sang list
        import json
        for sch in schedules_list:
            days_raw = sch['days_of_week']
            if isinstance(days_raw, str):
                try:
                    sch['days_of_week'] = json.loads(days_raw)
                except:
                    sch['days_of_week'] = []
            elif days_raw is None:
                sch['days_of_week'] = []

            # Chuyển date sang string để JSON serialize được
            if sch['start_date']:
                sch['start_date'] = str(sch['start_date'])
            if sch['end_date']:
                sch['end_date'] = str(sch['end_date'])

        # Lấy danh sách giáo viên cho dropdown
        cursor.execute("SELECT id, name, department FROM teachers ORDER BY name")
        teachers_list = cursor.fetchall()

       # Lấy danh sách lớp từ lịch học
        cursor.execute("""
                SELECT DISTINCT class_name
                FROM schedules
                WHERE class_name IS NOT NULL
                AND class_name <> ''
                ORDER BY class_name
            """)
        classes_list = [row['class_name'] for row in cursor.fetchall()]

        print("CLASSES_LIST =", classes_list)  # debug
        return render_template('manage_subjects.html',
                               subjects=subjects_list,
                               schedules=schedules_list,
                               teachers=teachers_list,
                               classes=classes_list)
    finally:
        cursor.close()
        db.close()


# API: Thêm môn học
@admin_bp.route('/api/subjects', methods=['POST'])
def api_add_subject():
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("""
            INSERT INTO subjects (name, class_name, teacher_id)
            VALUES (%s, %s, %s)
        """, (data['name'], data.get('class_name', ''), data.get('teacher_id') or None))
        db.commit()
        return jsonify({"id": cursor.lastrowid, "message": "Thêm môn học thành công"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


# API: Sửa môn học
@admin_bp.route('/api/subjects/<int:subject_id>', methods=['PUT'])
def api_edit_subject(subject_id):
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("""
            UPDATE subjects SET name=%s, class_name=%s, teacher_id=%s
            WHERE id=%s
        """, (data['name'], data.get('class_name', ''), data.get('teacher_id') or None, subject_id))
        db.commit()
        return jsonify({"message": "Cập nhật thành công"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


# API: Xóa môn học
@admin_bp.route('/api/subjects/<int:subject_id>', methods=['DELETE'])
def api_delete_subject(subject_id):
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("DELETE FROM schedules WHERE subject_id = %s", (subject_id,))
        cursor.execute("DELETE FROM subjects WHERE id = %s", (subject_id,))
        db.commit()
        return jsonify({"message": "Đã xóa môn học"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


# API: Thêm lịch học
@admin_bp.route('/api/schedules', methods=['POST'])
def api_add_schedule():
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401
    import json
    data = request.get_json()
    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        # Lấy class_name từ subject nếu không truyền lên
        cursor.execute("SELECT class_name FROM subjects WHERE id=%s", (data['subject_id'],))
        subj = cursor.fetchone()
        class_name = data.get('class_name') or (subj['class_name'] if subj else '')

        cursor.execute("""
            INSERT INTO schedules
                (subject_id, class_name, room, start_time, end_time,
                 start_date, end_date, days_of_week)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data['subject_id'],
            class_name,
            data.get('room', ''),
            data['start_time'],
            data['end_time'],
            data.get('start_date') or None,
            data.get('end_date')   or None,
            json.dumps(data.get('days_of_week', []))
        ))
        db.commit()
        return jsonify({"id": cursor.lastrowid, "message": "Thêm lịch học thành công"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


# API: Xóa lịch học
@admin_bp.route('/api/schedules/<int:schedule_id>', methods=['DELETE'])
def api_delete_schedule(schedule_id):
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("DELETE FROM schedules WHERE id=%s", (schedule_id,))
        db.commit()
        return jsonify({"message": "Đã xóa lịch học"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()

@admin_bp.route('/save_teacher', methods=['POST'])
def save_teacher():

    if session.get("role") != "admin":
        return jsonify({"success": False})

    data = request.get_json()

    teacher_id = data.get('id')
    name = data.get('name')
    department = data.get('department')
    email = data.get('email')
    status = data.get('status')

    db = get_db()
    cursor = db.cursor()

    try:

        # update teacher
        cursor.execute("""
            UPDATE teachers
            SET name=%s,
                department=%s,
                email=%s
            WHERE id=%s
        """, (name, department, email, teacher_id))

        # update user status
        cursor.execute("""
            UPDATE users u
            JOIN teachers t ON u.id = t.user_id
            SET u.status=%s
            WHERE t.id=%s
        """, (status, teacher_id))

        db.commit()

        return jsonify({
            "success": True
        })

    except Exception as e:
        db.rollback()
        print(e)

        return jsonify({
            "success": False,
            "error": str(e)
        })

    finally:
        cursor.close()
        db.close()


@admin_bp.route('/api/teacher/<int:teacher_id>/assign', methods=['POST'])
def save_teacher_assign(teacher_id):
    if session.get("role") != "admin":
        return jsonify({"success": False}), 401

    import json
    data     = request.get_json()
    subjects = data.get("subjects", [])

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        for s in subjects:
            name       = (s.get("name") or "").strip()
            class_name = s.get("class_name", "")
            subject_id = s.get("subject_id")

            if not name:
                continue

            # Upsert subject
            if subject_id:
                cursor.execute("""
                    UPDATE subjects SET name=%s, class_name=%s
                    WHERE id=%s AND teacher_id=%s
                """, (name, class_name, subject_id, teacher_id))
            else:
                cursor.execute("""
                    SELECT id FROM subjects
                    WHERE name=%s AND teacher_id=%s LIMIT 1
                """, (name, teacher_id))
                existing = cursor.fetchone()
                if existing:
                    subject_id = existing['id']
                    cursor.execute("""
                        UPDATE subjects SET class_name=%s WHERE id=%s
                    """, (class_name, subject_id))
                else:
                    cursor.execute("""
                        INSERT INTO subjects (name, teacher_id, class_name)
                        VALUES (%s, %s, %s)
                    """, (name, teacher_id, class_name))
                    subject_id = cursor.lastrowid

            # Xóa lịch cũ rồi insert lại
            cursor.execute("DELETE FROM schedules WHERE subject_id=%s", (subject_id,))

            for sc in s.get("schedules", []):
                days = sc.get("days_of_week", [])
                if not sc.get("start_time") or not sc.get("end_time"):
                    continue
                cursor.execute("""
                    INSERT INTO schedules
                        (subject_id, class_name, room,
                         start_time, end_time,
                         start_date, end_date, days_of_week)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    subject_id,
                    class_name,
                    sc.get("room", ""),
                    sc["start_time"],
                    sc["end_time"],
                    sc.get("start_date") or None,
                    sc.get("end_date")   or None,
                    json.dumps(days)
                ))

        db.commit()
        return jsonify({"success": True})

    except Exception as e:
        db.rollback()
        print("🔥 save_teacher_assign error:", e)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@admin_bp.route('/api/teacher/<int:teacher_id>/assign/<int:subject_id>', methods=['DELETE'])
def delete_teacher_subject(teacher_id, subject_id):
    if session.get("role") != "admin":
        return jsonify({"success": False}), 401

    db = get_db()
    cursor = db.cursor()
    try:
        # Xóa lịch học liên quan trước
        cursor.execute("DELETE FROM schedules WHERE subject_id = %s", (subject_id,))
        # Xóa môn học
        cursor.execute("""
            DELETE FROM subjects
            WHERE id = %s AND teacher_id = %s
        """, (subject_id, teacher_id))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cursor.close()
        db.close()
    
@admin_bp.route('/api/check_schedule_conflict', methods=['POST'])
def check_schedule_conflict():
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401

    data        = request.get_json()
    teacher_id  = data.get('teacher_id')
    schedules   = data.get('schedules', [])

    db = get_db()
    cursor = db.cursor(dictionary=True)
    conflicts = []

    try:
        for sc in schedules:
            class_name  = sc.get('class_name', '').strip()
            days        = sc.get('days_of_week', [])
            start_time  = sc.get('start_time', '')
            end_time    = sc.get('end_time', '')
            exclude_ids = [int(x) for x in sc.get('exclude_schedule_ids', []) if x]

            if not class_name or not days or not start_time or not end_time:
                continue

            # Đảm bảo days là list số nguyên
            days = [int(d) for d in days]

            # Đảm bảo format HH:MM:SS
            st = start_time.strip() + ':00' if len(start_time.strip()) == 5 else start_time.strip()
            et = end_time.strip()   + ':00' if len(end_time.strip())   == 5 else end_time.strip()

            # placeholder: nếu không có id cần loại trừ thì dùng -1 (không bao giờ tồn tại)
            excl_placeholder = ','.join(['%s'] * len(exclude_ids)) if exclude_ids else '-1'
            excl_params      = exclude_ids if exclude_ids else []

            for day in days:
                day_json = json.dumps(int(day))  # '5' kiểu JSON number, đúng cho JSON_CONTAINS

                # ── Trùng lớp ──────────────────────────────────────────────
                q_class = f"""
                    SELECT sch.id, sch.class_name, sch.start_time, sch.end_time,
                           sub.name AS subject_name, t.name AS teacher_name
                    FROM schedules sch
                    JOIN subjects sub ON sch.subject_id = sub.id
                    LEFT JOIN teachers t ON sub.teacher_id = t.id
                    WHERE sch.class_name = %s
                      AND sch.id NOT IN ({excl_placeholder})
                      AND JSON_CONTAINS(sch.days_of_week, %s, '$')
                      AND sch.start_time < %s
                      AND sch.end_time   > %s
                """
                cursor.execute(q_class, [class_name] + excl_params + [day_json, et, st])
                for row in cursor.fetchall():
                    conflicts.append({
                        'type': 'class',
                        'message': (
                            f"Lớp {class_name} đã có lịch môn '{row['subject_name']}' "
                            f"(GV: {row['teacher_name']}) vào thứ {day} "
                            f"{td_to_str(row['start_time'])}–{td_to_str(row['end_time'])}"
                        ),
                    })
# ── Trùng giáo viên ────────────────────────────────────────
                if teacher_id:
                    q_teacher = f"""
                        SELECT sch.id, sch.class_name, sch.start_time, sch.end_time,
                               sub.name AS subject_name
                        FROM schedules sch
                        JOIN subjects sub ON sch.subject_id = sub.id
                        WHERE sub.teacher_id = %s
                          AND sch.id NOT IN ({excl_placeholder})
                          AND JSON_CONTAINS(sch.days_of_week, %s, '$')
                          AND sch.start_time < %s
                          AND sch.end_time   > %s
                    """
                    cursor.execute(q_teacher, [teacher_id] + excl_params + [day_json, et, st])
                    for row in cursor.fetchall():
                        conflicts.append({
                            'type': 'teacher',
                            'message': (
                                f"Giáo viên đã có lịch dạy lớp {row['class_name']} "
                                f"môn '{row['subject_name']}' vào thứ {day} "
                                f"{td_to_str(row['start_time'])}–{td_to_str(row['end_time'])}"
                            ),
                        })

        return jsonify({"conflicts": conflicts})
    except Exception as e:
        print("check_conflict error:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@admin_bp.route('/export_timetable')
def export_timetable():
    if session.get("role") != "admin":
        return redirect("/login")

    class_name = request.args.get("class_name")

    if not class_name:
        return "Thiếu class_name", 400

    from io import BytesIO
    import json

    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT
                sch.class_name,
                sch.room,
                TIME_FORMAT(sch.start_time,'%H:%i') AS start_time,
                TIME_FORMAT(sch.end_time,'%H:%i') AS end_time,
                sch.days_of_week,
                s.name AS subject_name,
                t.name AS teacher_name
            FROM schedules sch
            JOIN subjects s
                ON sch.subject_id = s.id
            LEFT JOIN teachers t
                ON s.teacher_id = t.id
            WHERE sch.class_name = %s
            ORDER BY sch.start_time
        """, (class_name,))

        schedules = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = class_name

        # ==================================================
        # TIÊU ĐỀ
        # ==================================================

        ws.merge_cells("A1:G1")

        ws["A1"] = f"THỜI KHÓA BIỂU LỚP {class_name}"

        ws["A1"].font = Font(
            size=16,
            bold=True
        )

        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # ==================================================
        # HEADER
        # ==================================================

        headers = [
            "Tiết / Ngày",
            "Thứ 2",
            "Thứ 3",
            "Thứ 4",
            "Thứ 5",
            "Thứ 6",
            "Thứ 7"
        ]

        for col, text in enumerate(headers, start=1):

            cell = ws.cell(
                row=3,
                column=col
            )

            cell.value = text

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="4F46E5"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # ==================================================
        # 5 KHUNG GIỜ CỐ ĐỊNH
        # ==================================================

        time_slots = [
            ("Tiết 1-2\n07:00-08:30", 4),
            ("Tiết 3-4\n08:45-10:15", 5),
            ("Tiết 5-6\n10:30-12:00", 6),
            ("Tiết 7-8\n13:00-14:30", 7),
            ("Tiết 9-10\n14:45-16:15", 8)
        ]

        for label, row in time_slots:

            cell = ws.cell(
                row=row,
                column=1
            )

            cell.value = label

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # ==================================================
        # MAP THỨ
        # ==================================================

        day_column_map = {
            1: 2,  # Thứ 2
            2: 3,  # Thứ 3
            3: 4,  # Thứ 4
            4: 5,  # Thứ 5
            5: 6,  # Thứ 6
            6: 7   # Thứ 7
        }

        # ==================================================
        # XÁC ĐỊNH DÒNG THEO GIỜ HỌC
        # ==================================================

        def get_row_by_time(start_time):

            start_time = str(start_time)[:5]

            if start_time < "08:45":
                return 4      # Tiết 1-2

            elif start_time < "10:30":
                return 5      # Tiết 3-4

            elif start_time < "13:00":
                return 6      # Tiết 5-6

            elif start_time < "14:45":
                return 7      # Tiết 7-8

            else:
                return 8      # Tiết 9-10

        # ==================================================
        # ĐỔ DỮ LIỆU
        # ==================================================

        for item in schedules:

            days_raw = item["days_of_week"]

            if isinstance(days_raw, str):
                try:
                    days = json.loads(days_raw)
                except Exception:
                    days = []
            else:
                days = days_raw or []

            row_excel = get_row_by_time(
                item["start_time"]
            )

            content = (
                f"{item['subject_name']}\n"
                f"GV: {item['teacher_name'] or ''}\n"
                f"Phòng: {item['room'] or ''}"
            )

            for day in days:

                try:
                    day = int(day)
                except:
                    continue

                col_excel = day_column_map.get(day)

                if not col_excel:
                    continue

                cell = ws.cell(
                    row=row_excel,
                    column=col_excel
                )

                cell.value = content

                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

        # ==================================================
        # STYLE
        # ==================================================

        thin = Side(style="thin")

        for row in ws.iter_rows(
            min_row=3,
            max_row=8,
            min_col=1,
            max_col=7
        ):
            for cell in row:
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

        ws.column_dimensions["A"].width = 22

        for col in ["B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 35

        for r in range(4, 9):
            ws.row_dimensions[r].height = 80

        ws.freeze_panes = "B4"

        # ==================================================
        # XUẤT FILE
        # ==================================================

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"thoi_khoa_bieu_{class_name}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    finally:
        cursor.close()
        db.close()

